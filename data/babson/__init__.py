import datetime
import itertools
import pathlib
import re
from collections import defaultdict

import pytz
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework import serializers

from backend.models import Ingredient, MealItem, MealSelection, School
from data.babson.parse_items_excel import get_meal_items

SCHOOL_ID = 10
VERSION = 0
FILE_DIR = pathlib.Path(__file__).resolve().parent

MEAL_TIMES = (
    ('breakfast', datetime.time(hour=7, minute=30)),
    ('lunch', datetime.time(hour=11)),
    ('afterlunch', datetime.time(hour=14, minute=30)),
    ('dinner', datetime.time(hour=16, minute=30)),
)
MEAL_TIMES_DICT = dict(MEAL_TIMES)

MEAL_KEYS = tuple(map(lambda x: x[0], MEAL_TIMES))
MEAL_SELECTION_COLS = [0, 1, 4, 5, 6, 7, 10]

ALLERGY_INGREDIENTS = ['peanuts', 'tree_nuts', 'eggs', 'soy', 'wheat', 'fish', 'shellfish', 'corn', 'gelatin']


def clean_old():
    MealSelection.objects.filter(school__id=SCHOOL_ID, version=0).delete()
    MealItem.objects.filter(school__id=SCHOOL_ID, version=0).delete()
    Ingredient.objects.filter(school__id=SCHOOL_ID, version=0).delete()


def add_ingredients():
    babson = School.objects.get(pk=SCHOOL_ID)
    objs = []
    for name in ALLERGY_INGREDIENTS:
        objs.append(Ingredient(name=name, school=babson, version=VERSION))
    objs = Ingredient.objects.bulk_create(objs)
    return {obj.name: obj for obj in objs}


def add_meal_items():
    meal_items, _ = get_meal_items()

    class MealItemSerializer(serializers.ModelSerializer):
        class Meta:
            model = MealItem
            exclude = ['ingredients']

    meals = []
    trim_ids = []
    objs = []
    for m_info in meal_items.values():
        ser = MealItemSerializer(data=m_info)
        ser.is_valid(raise_exception=True)
        objs.append(MealItem(**ser.validated_data))
        meals.append(str(m_info['meal']))
        trim_ids.append(str(m_info['trim_id']))

    return list(zip(MealItem.objects.bulk_create(objs), meals, trim_ids))


# input is a list of tuples (meal_item, meal) (meal = 'breakfast', 'lunch', 'dinner'
def add_meals(meal_items: list[tuple[MealItem, str, str]]):
    # setup item dicts from meal items input
    items_by_meal = defaultdict(list)
    for item, meal, trim_id in meal_items:
        items_by_meal[meal].append((item, trim_id))
    items_by_meal['afterlunch'] = \
        list(filter(lambda x: x[0].station in ['FLAME', '500 Degrees', 'Carved and Crafted'], items_by_meal['lunch']))
    # Need to setup afterlunch meal
    items_by_id: dict[str, list[tuple[MealItem, str]]] = defaultdict(list)
    for meal, items_and_ids in items_by_meal.items():
        for item, trim_id in items_and_ids:
            items_by_id[trim_id].append((item, meal))

    # parse meals sheet
    meal_sheet: Worksheet = load_workbook(FILE_DIR / 'Menuworks_.xlsx', read_only=True)['Report']

    item_ids_by_day: dict[datetime.date, list[str]] = defaultdict(list)
    meals: list[MealSelection] = []
    corr_items: list[list[MealItem]] = []
    items_to_update: list[MealItem] = []
    babson = School.objects.get(pk=SCHOOL_ID)

    for col in MEAL_SELECTION_COLS:
        meal_day = None
        item_name = None
        for row in meal_sheet.iter_rows(13, 1481):
            val = row[col].value
            if val is None:
                continue
            elif m := re.match(r'\w+ \((\d+)/(\d+)/(\d+)\)', val):
                meal_day = datetime.date(year=int(m.group(3)), month=int(m.group(1)), day=int(m.group(2)))
            elif val in items_by_id:
                item_ids_by_day[meal_day].append(val)
                for item, _ in items_by_id[val]:
                    item.name = item_name
                    items_to_update.append(item)
            else:
                item_name = val
    print('Parsed spreadsheet')

    # Create objects
    for day, item_ids in item_ids_by_day.items():
        _i = itertools.chain(*(items_by_id[item_id] for item_id in item_ids))
        _snd = lambda t: t[1]

        for meal, items in itertools.groupby(sorted(_i, key=_snd), key=_snd):
            # The Waterloo student did this one
            dt = datetime.datetime.combine(day, MEAL_TIMES_DICT[meal], pytz.timezone('America/Toronto'))
            m = MealSelection(name=f'{meal.title()} on {day.strftime("%A, %B")} {day.day}, {day.year}',
                              group=meal,
                              timestamp=dt,
                              school=babson,
                              version=VERSION)
            meals.append(m)
            corr_items.append(list(map(lambda x: x[0], items)))
    print('Created Meal Selection Objects')

    meals = MealSelection.objects.bulk_create(meals)
    print(f'Created {len(meals)} meal objects')
    for idx, (m, i) in enumerate(zip(meals, corr_items), start=1):
        m.items.set(i)
        m.save()

        if idx % 40 == 0:
            print(f'Updated `items` field for {idx} meal objects')

    MealItem.objects.bulk_update(items_to_update, fields=('name',))
    print(f'Renamed {len(items_to_update)} items')


def setup():
    clean_old()
    print('Cleaned old data')
    _ = add_ingredients()
    print(f'Added {len(_)} ingredients')
    meal_items = add_meal_items()
    print(f'Added {len(meal_items)} meal items to DB')
    add_meals(meal_items)
    print('Setup complete!')
