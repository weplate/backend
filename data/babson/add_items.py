import json
import pathlib
import re
from re import Match

from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from rest_framework import serializers

from backend.models import MealItem

OUT_FILE_PATH = 'meal_items.json'
SCHOOL_ID = 10
MAX_NAME_LEN = 64
FILE_DIR = pathlib.Path(__file__).resolve().parent


def parse_meal_items(version):
    main_sheet: Worksheet = load_workbook(FILE_DIR / 'Nutrition_Facts_as_of_Feb_20_1.xlsx', read_only=True)['Report']
    micros_sheet: Worksheet = load_workbook(FILE_DIR / 'Additional_nutrition_facts_as_of_Feb_20.xlsx', read_only=True)['Report']

    meal_items = {}

    NUM = r'(([0-9]+)(\/[0-9]+)?)'

    def parse_num(matches: Match):
        mat = matches.group(1)
        if '/' in mat:
            n, d = mat.split('/')
            return float(n) / float(d)
        else:
            return float(mat)

    bad_units = set()

    def parse_portion(value):
        if m := re.search(rf'{NUM} cup', value):
            return parse_num(m) * 236.588, False
        elif m := re.search(rf'{NUM} pint', value):
            return parse_num(m) * 473, False
        elif m := re.search(rf'{NUM} tbsp', value):
            return parse_num(m) * 14.7866, False
        elif m := re.search(rf'{NUM} tsp', value):
            return parse_num(m) * 4.92892, False
        elif m := re.search(rf'{NUM} floz', value):
            return parse_num(m) * 29.5735, False
        elif m := re.search(rf'[0-9] ladle-{NUM}oz', value):
            return parse_num(m) * 29.5735, False
        else:
            bad_units.add(value)
            return -1, True

    # Read macros
    cur_station = None
    unparseable = 0
    for row in main_sheet.iter_rows(13, 1039):
        def get_col(col):
            return row[col].value

        def num_col(col):
            val = get_col(col)
            if val is None:
                return 0
            else:
                only_num = re.sub(r'[^0-9.]', '', val)
                return float(only_num) if only_num else 0

        if get_col(2) is None:
            cur_station = get_col(0)
        else:
            portion, err = parse_portion(get_col(4))

            station_meal, station_name = cur_station.split(' - ', 1)
            station_meal = station_meal.lower()

            if station_name.lower() not in ['homestyle', 'rooted', 'fyul', 'flame', '500 degrees', 'carved and crafted']:
                err = True
            else:
                if err:
                    unparseable += 1

            if not err:
                name = re.sub(r'(CHE( (\d+))? (- )?)|(HC )|([\w+]+: )',
                               '', get_col(0), count=1)[:MAX_NAME_LEN]
                meal_items[get_col(0)] = {
                    # basic info
                    'name': name,
                    'school': SCHOOL_ID,
                    'version': version,
                    'station': station_name,

                    # basic nutritional/number info
                    'portion_weight': num_col(7),
                    'portion_volume': portion,
                    'ingredients': [],
                    'category': c.lower() if (c := get_col(1)) else None,
                    'cafeteria_id': get_col(2),

                    # sheet 1 nutrients
                    'calories': num_col(8),
                    'protein': num_col(10),
                    'total_fat': num_col(11),
                    'saturated_fat': num_col(12),
                    'carbohydrate': num_col(13),
                    'sugar': num_col(14),
                    'fiber': num_col(16),
                    'sodium': num_col(18),
                    'potassium': num_col(19),
                    'calcium': num_col(20),
                    'iron': num_col(22),

                    # Extra property for other parsers
                    'meal': station_meal,
                }

    for row in micros_sheet.iter_rows(13, 1021):
        def get_col(col):
            return row[col].value

        def num_col(col):
            val = get_col(col)
            if val is None:
                return 0
            else:
                only_num = re.sub(r'[^0-9.]', '', val)
                return float(only_num) if only_num else 0

        if get_col(0) in meal_items:
            meal_items[get_col(0)] |= {
                'vitamin_d': num_col(7),
                'vitamin_c': num_col(9),
                'vitamin_a': num_col(10),
                'cholesterol': num_col(13),
            }

    print(f'Parsed {len(meal_items)} meal items')
    print(f'Got {unparseable} items with correct stations but otherwise unreadable')

    return meal_items, bad_units


def add_meal_items(version):
    meal_items, _ = parse_meal_items(version)

    class MealItemSerializer(serializers.ModelSerializer):
        class Meta:
            model = MealItem
            exclude = ['ingredients']

    meals = []
    trim_ids = []
    objs = []
    for m_info in meal_items.values():
        ser = MealItemSerializer(data=m_info)
        ser.is_valid(raise_exception=True)
        objs.append(MealItem(**ser.validated_data))
        meals.append(str(m_info['meal']))
        trim_ids.append(str(m_info['cafeteria_id']))

    return list(zip(MealItem.objects.bulk_create(objs), meals, trim_ids))


def main():
    meal_items, bad_units = parse_meal_items(-1)

    with open(OUT_FILE_PATH, 'w') as f:
        json.dump(meal_items, f)

    print('-- Bad Units --')
    for unit in sorted(bad_units):
        print(unit)


if __name__ == '__main__':
    main()